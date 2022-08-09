import time, sys, unittest, random, json, openpyxl, platform
from datetime import datetime
from appium import webdriver
from random import randint
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.alert import Alert

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell

# Connect to Appium with the below desire capabilities
# http://appium.io/docs/en/writing-running-appium/caps/
dc = {
    "deviceName": "R58MC3QBSZT",
    "platformName": "Android",
    "app": "E:\\Quynh\\hr-app\\hanbiro-global-hr-app-v2.1.7.apk",
    "automationName": "UiAutomator2",
    "autoGrantPermissions": "true",
    "appWaitPackage": "com.hanbiro.hanbirohrm",
    "appWaitActivity": "com.hanbiro.globalhr.MainActivity",
    "locale": "US",
    "language": "en"
}

with open("NQ_HR_config.json") as json_data_file:
    data = json.load(json_data_file)

n = random.randint(1,3000)

# If desire capabilities are valid, the app will be open at Log in screen
driver = webdriver.Remote('http://localhost:4723/wd/hub', dc)

class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
    testcase_pass = "Test case status: pass"
    testcase_fail = "Test case status: fail"

if platform == "linux" or platform == "linux2":
    local_path = "/home/oem/groupware-auto-test"
    json_file = local_path + "/NQ_HR_config.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    log_folder = "/Log/"
    log_testcase = "/Log/"
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("execution_log_", "fail_log_")
    error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "NQuynh_Testcase_HRApp_" + str(objects.date_id) + ".xlsx"
else :
    local_path = "E:\\Quynh\\hr-app"
    json_file = local_path + "\\NQ_HR_config.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    log_folder = "\\Log\\"
    log_testcase = "\\Log\\"
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("execution_log_", "fail_log_")
    error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "NQuynh_Testcase_HRApp_" + str(objects.date_id) + ".xlsx"

logs = [execution_log, fail_log, error_log, testcase_log]
for log in logs:
    if ".txt" in log:
        open(log, "x").close()
    else:
        wb = Workbook()
        myFill = PatternFill(start_color='adc5e7',
                   end_color='adc5e7',
                   fill_type='solid',)
        font = Font(name='Calibri',
                    size=11 ,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        ws = wb.active

        ws.cell(row=1, column=1).value= "Menu"
        ws.cell(row=1, column=2).value = "Sub-Menu"
        ws.cell(row=1, column=3).value = "Test Case Name"
        ws.cell(row=1, column=4).value = "Status"
        ws.cell(row=1, column=5).value = "Description"
        ws.cell(row=1, column=6).value = "Date"
        ws.cell(row=1, column=7).value = "Tester"
        # color 
        ws.cell(row=1, column=1).fill = myFill
        ws.cell(row=1, column=2).fill = myFill
        ws.cell(row=1, column=3).fill = myFill
        ws.cell(row=1, column=4).fill = myFill
        ws.cell(row=1, column=5).fill = myFill
        ws.cell(row=1, column=6).fill = myFill
        ws.cell(row=1, column=7).fill = myFill
        # font
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        ws.cell(row=1, column=3).font = Font(bold=True)
        ws.cell(row=1, column=4).font = Font(bold=True)
        ws.cell(row=1, column=5).font = Font(bold=True)
        ws.cell(row=1, column=6).font = Font(bold=True)
        ws.cell(row=1, column=7).font = Font(bold=True)

        wb.save(log)

def Logging(msg):
    Logging(msg)
    log_msg = open(execution_log, "a")
    log_msg.write(str(msg) + "\n")
    log_msg.close()

def TesCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    Logging(description)

    # if status == "Pass":
    #     Logging(objects.testcase_pass)
    # else:
    #     Logging(objects.testcase_fail)

    wb = openpyxl.load_workbook(testcase_log)
    current_sheet = wb.active
    start_row = len(list(current_sheet.rows)) + 1

    current_sheet.cell(row=start_row, column=1).value = menu
    current_sheet.cell(row=start_row, column=2).value = sub_menu
    current_sheet.cell(row=start_row, column=3).value = testcase
    current_sheet.cell(row=start_row, column=4).value = status
    current_sheet.cell(row=start_row, column=5).value = description
    current_sheet.cell(row=start_row, column=6).value = objects.date_time
    current_sheet.cell(row=start_row, column=7).value = tester

    # Apply color for status: Pass/Fail
    passFill = PatternFill(start_color='b6d7a8',
                   end_color='b6d7a8',
                   fill_type='solid',)
    failFill = PatternFill(start_color='ea9999',
                   end_color='ea9999',
                   fill_type='solid')
    if status == "Pass":
        Logging(objects.testcase_pass)
        current_sheet.cell(row=start_row, column=4).fill = passFill
    else:
        Logging(objects.testcase_fail)
        current_sheet.cell(row=start_row, column=4).fill = failFill
    wb.save(testcase_log)

def ValidateFailResultAndSystem(fail_msg):
    Logging(fail_msg)
    append_fail_result = open(fail_log, "a")
    append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
    append_fail_result.close()

def execution():
    try:
        checkcrashapp = WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Domain']")))
        if checkcrashapp.is_displayed():
            Logging("------- Login to app -------")
            # Input information for log-in
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Domain']")))
            driver.find_element_by_xpath(data["domain"]).send_keys(data["domain_name"])
            Logging("- Input Domain")
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='ID']")))
            driver.find_element_by_xpath(data["id_app"]).send_keys(data["id_name"])
            Logging("- Input ID")
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Password']")))
            driver.find_element_by_xpath(data["password"]).send_keys(data["pass_input"])
            Logging("- Input Password")
            WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Login')]")))
            driver.find_element_by_xpath(data["button_login"]).click()
            Logging("=> Click Log In button")
            driver.implicitly_wait(50)

            ''' * Check crash app done
            => Execute function '''
            # GPS()
            # break_time()
            # clock_out()
            # viewNoti()
            # add_event()
            # admin_GPS()
            # timecard()
            # settings()
            # vacation()
            # approve_request()
            check_notification()
        else:
            Logging("=> Crash app")
            exit(0)
    except WebDriverException:
            Logging("=> Crash app")
            exit(0)

def clock_in():
    Logging(" ")
    Logging("------- Clock In -------")
    try:
        title_app = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["title"])))
        if title_app.text == 'GPS':
            Logging("Clock in - GPS")
            try:
                OT = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["clock_in"]["nightshift"])))
                if OT.text == 'Night Shift':
                    Logging("=> Work night shift")
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["OT"]["confirm_OT"]))).click()
                    Logging("- Confirm OT / Work night shift")
                    time.sleep(5)
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["OT"]["OT_apply"]))).click()
                    Logging("- Apply OT")
                    time.sleep(5)
                    
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["OT"]["select_time"]))).click()
                    Logging("- Select time")
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["OT"]["time"]))).click()
                    time.sleep(5)
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please input a reason to apply OT')]"))).send_keys(data["OT"]["text"])
                    Logging("-> Input reason")

                    driver.swipe(start_x=675, start_y=2458, end_x=675, end_y=2000, duration=800)
                    time.sleep(5)
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["OT"]["apply_OT"]))).click()
                    time.sleep(5)

                    try:
                        check_OT = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["OT"]["apply_text"])))
                        if check_OT.text == 'Apply OT':
                            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Close')]"))).click()
                            Logging("=> Apply OT success")
                            TesCase_LogResult(**data["testcase_result"]["timecard"]["apply_OT_GPS"]["pass"])
                        else:
                            Logging("=> Crash app")
                            TesCase_LogResult(**data["testcase_result"]["timecard"]["apply_OT_GPS"]["fail"])
                    except WebDriverException:
                        Logging("=> Crash app")     
                        TesCase_LogResult(**data["testcase_result"]["timecard"]["apply_OT_GPS"]["fail"])   
                        exit(0)
                else:
                    Logging("=> Apply OT not display")
            except WebDriverException:
                Logging("=> Apply OT not display")  
        
            try:
                icon_clock_in = driver.find_element_by_xpath(data["clock_in"]["icon_clock_in_button"])
                if icon_clock_in.is_displayed():
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,data["clock_in"]["icon_clockin"]))).click()
                    Logging("=> Clock In with GPS")
                    try:
                        status_late = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["clock_in_late"])))
                        if status_late.text == 'Tardiness':
                            Logging("=> Clock in late")
                            driver.find_element_by_xpath(data["clock_in"]["text_input"]).send_keys(data["clock_in"]["text"])
                            Logging("- Input reason")
                            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Save')]")))
                            driver.find_element_by_xpath(data["clock_in"]["save_button"]).click()
                            Logging("=> Save")
                            TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_in_GPS"]["pass"])
                        else:
                            Logging("=> Clock in on time") 
                            TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_in_GPS"]["pass"])
                            time.sleep(5)
                    except WebDriverException:
                        Logging("=> Crash app")
                        TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_in_GPS"]["fail"])
                        exit(0)
                else:
                    Logging(" Clock in button not display")
                    time.sleep(5)
            except WebDriverException:
                Logging("=> Clock in button not display")
            time.sleep(5)

            try:
                icon_clock_out = driver.find_element_by_xpath(data["clock_in"]["icon_breaktime"])
                if icon_clock_out.is_displayed():
                    Logging("=> Already clock in")
                    TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_in_GPS"]["pass"])
                else:
                    Logging("=> Break time button not display")
                    TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_in_GPS"]["fail"])
            except WebDriverException:
                Logging("=> Fail") 
                TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_in_GPS"]["fail"])
        else:
            Logging("=> .Crash app") 
            TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_in_GPS"]["fail"])
            time.sleep(5)
        
            

        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["preview"]))).click()
        Logging("- View preview date")
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["next"]))).click()
        Logging("- View next date")
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["calendar"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='0']//android.view.ViewGroup[@index='3']//android.widget.ScrollView[@index='0']//android.view.ViewGroup[@index='1']//android.widget.Button[@index=19]"))).click()
        Logging("- Select date")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["select"]))).click()
    except WebDriverException:
        Logging("=> Crash app")
        exit(0)

def GPS():
    clock_in()
    try:
        try:
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["IN"]))).click()
            Logging("- Select IN on map")
            time.sleep(5)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["select_branch"]))).click()
            Logging("- Select branch")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["close_popup"]))).click()
            Logging("- Close popup")
            time.sleep(5)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["OUT"]))).click()
            Logging("- Select OUT on map")
            time.sleep(5)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["select_branch"]))).click()
            Logging("- Select branch")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["close_popup"]))).click()
            Logging("- Close popup")
            time.sleep(5)
            TesCase_LogResult(**data["testcase_result"]["timecard"]["branch_name"]["pass"])
        except WebDriverException:
            Logging("=> IN - OUT not display")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Wifi')]"))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'GPS')]"))).click() 
            Logging("- Change to wifi -> GPS") 
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["IN"]))).click()
            Logging("- Select IN on map")
            time.sleep(5)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["select_branch"]))).click()
            Logging("- Select branch")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["close_popup"]))).click()
            Logging("- Close popup")
            time.sleep(5)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["OUT"]))).click()
            Logging("- Select OUT on map")
            time.sleep(5)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["select_branch"]))).click()
            Logging("- Select branch")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_in"]["close_popup"]))).click()
            Logging("- Close popup")
            TesCase_LogResult(**data["testcase_result"]["timecard"]["branch_name"]["pass"])
    except:
        TesCase_LogResult(**data["testcase_result"]["timecard"]["noti"]["fail"])
        Logging("- Can't select branch name")

def viewNoti():
    Logging(" ")
    Logging("------- View notification -------")
    try:
        time.sleep(10)
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["view_noti"]["noti"]))).click()
        Logging("=> Click view notification")
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["view_noti"]["back_homepage"]))).click()
        Logging("=> Back to homepage")       
        time.sleep(5)
        TesCase_LogResult(**data["testcase_result"]["timecard"]["noti"]["pass"])
    except:
        Logging("- Can't view noti")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["noti"]["fail"])
    
def break_time():
    ''' Break time button '''
    try:
        Logging("** Check break time - clock out")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["admin"]["homepage"]))).click()
        text_breaktime = driver.find_element_by_xpath(data["clock_in"]["breaktime_text"])
        if text_breaktime.text == 'Break Time':
            Logging("=> Start break time")
            driver.find_element_by_xpath(data["clock_in"]["icon_breaktime"]).click()
            time.sleep(30)
            ''' End break time '''
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'END BREAK TIME')]")))
            driver.find_element_by_xpath(data["clock_in"]["end_break_time"]).click()
            time.sleep(10)
            Logging("=> End Break time")
        else: 
            Logging("=> Already clock out")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["admin"]["homepage"]))).click()
        time.sleep(5)
    except WebDriverException:
        Logging("=> Already clock out") 
        time.sleep(5)

def clock_out():
    try:
        Logging("** Check clock out")
        text_breaktime = driver.find_element_by_xpath(data["clock_in"]["breaktime_text"])
        if text_breaktime.text == 'Break Time':
            clockout = driver.find_element_by_xpath(data["clock_out"]["icon_clock_out_button"])
            if clockout.is_displayed():
                clockout.click()
                Logging("=> Click clock out")
                time.sleep(2)
                ''' Check clock out time '''
                popup = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_out"]["status"])))
                if popup.text == 'Leave Early':
                    Logging("- Clock out early")
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please enter your reason')]"))).send_keys(data["clock_in"]["text"])
                    Logging("- Input reason")
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_out"]["save"]))).click()
                    Logging("=> Save")
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["clock_out"]["close"]))).click()
                    Logging("- Close")
                    TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_out_GPS"]["pass"])
                else:
                    Logging("=> Clock out on time")
                    driver.find_element_by_xpath(data["clock_out"]["close_popup"]).click()
                    TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_out_GPS"]["pass"])
        else:
            Logging("=> Already clock out")
            TesCase_LogResult(**data["testcase_result"]["timecard"]["clock_out_GPS"]["pass"])
    except WebDriverException:
        Logging("=> Already clock out") 
        time.sleep(5)

def clock_out_WF():
    try:
        Logging("** Check clock out")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Wifi')]"))).click()
        Logging("=> Change to wifi")
        time.sleep(10)
        clock_out()
    except WebDriverException:
        Logging("=> Already clock out") 
    time.sleep(5)

def clock_out_BC():
    try:
        Logging("** Check clock out")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Beacon')]"))).click()
        Logging("=> Change to Beacon")
        clock_out()
    except WebDriverException:
        Logging("=> Already clock out") 
        time.sleep(5)

def add_event():
    try:
        Logging(" ")
        Logging("------- Add event -------")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["timecard"]))).click()
        Logging("- Select time card")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["timesheet"]))).click()
        Logging("- Select time sheet")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["add"]))).click()
        Logging("- Select add")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please input data.')]"))).send_keys(data["event"]["title_text"])
        Logging("- Input title")
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["choose_event"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["type_event"]))).click()
        Logging("- Choose event")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["select_color"]))).click()
        Logging("- Select color")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please input data.')]"))).send_keys(data["event"]["location_text"])
        Logging("- Input location")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please add a memo.')]"))).send_keys(data["event"]["memo_text"])
        Logging("- Input memo")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["button_save"]))).click()
        TesCase_LogResult(**data["testcase_result"]["timecard"]["event"]["pass"])
    except:
        Logging("- Can't create event")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["event"]["fail"])

    Logging("** Check event use approval type")
    try:
        approval_type = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["popup"])))
        if approval_type.text == '[Approved] Your request approval request has been approved automatically':
            Logging("=> Use approval type: Automatic approval")

        elif approval_type.text == 'The approval request has been submitted. Please wait until the approval is completed.':
            Logging("=> Use approval type: Approval Line")

        elif approval_type.text == 'The approval request has been delivered to Head of Department. Please wait until the approval is completed.':
            Logging("=> Use approval type: Head Dept.")

        elif approval_type.text == 'The approval request has been delivered to Timecard Managers. Please wait until the approval is completed.':
            Logging("=> Use approval type: Timecard Manager")
        else:
            Logging("=> Use approval type: Dept. Manager")
    except WebDriverException:
        Logging("=> Use approval type: Dept. Manager") 

    driver.find_element_by_xpath(data["event"]["close_popup"]).click()
    Logging("=> Save event")
    time.sleep(5)

def admin_GPS():
    Logging(" ")
    try:
        Logging("------- Add GPS -------")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["admin"]["Admin"]))).click()
        Logging("- Select admin")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'GPS Setting')]"))).click()
        Logging("- Click GPS Setting")
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["admin"]["GPS_settings"]["add_gps"]))).click()
        Logging("- Add GPS")

        try:
            check_gps = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["admin"]["GPS_settings"]["popup"])))
            if check_gps.is_displayed():
                check_gps.click()
            else:
                Logging("=> Crash app")
        except WebDriverException:
            Logging("=> Crash app")        
            exit(0)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please input a address')]")))
        address = driver.find_element_by_xpath(data["admin"]["GPS_settings"]["search"])
        address.click()
        address.send_keys(data["admin"]["GPS_settings"]["search_text"])
        Logging("- Search address")

        driver.hide_keyboard()
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["admin"]["GPS_settings"]["list"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'400 Nguyễn Thị Thập, Tân Quy, District 7, Ho Chi Minh City, Vietnam')]"))).click()
        Logging("- Select address")

        location = driver.find_element_by_xpath(data["admin"]["GPS_settings"]["location"])
        location.clear()
        location.send_keys(data["admin"]["GPS_settings"]["location_text"] + str(n))
        Logging("- Input location")
        driver.find_element_by_xpath(data["admin"]["GPS_settings"]["workPlace"]).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'VietNam')]"))).click()
        #driver.find_element_by_xpath(data["admin"]["GPS_settings"]["workPlace_input"]).click()
        Logging("- Select work Place")

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Save')]"))).click()
        driver.find_element_by_xpath(data["admin"]["GPS_settings"]["close_save"]).click()
        Logging("- Save")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["add_GPS"]["pass"])

        time.sleep(5)
    except:
        Logging("- Can't add GPS")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["add_GPS"]["fail"])

    try:
        Logging("- Search GPS")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please enter your keyword.')]"))).click()
        
        ''' Send key "hr-GPS" from keyboard mobile '''
        driver.is_keyboard_shown()
        driver.press_keycode(36)
        driver.press_keycode(46)
        driver.press_keycode(69)
        driver.press_keycode(35)
        driver.press_keycode(44)
        driver.press_keycode(47)
        driver.press_keycode(66)

        Logging("** Edit GPS")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'hr-GPS')]"))).click()
        location = driver.find_element_by_xpath(data["admin"]["GPS_settings"]["location"])
        location.clear()
        location.send_keys(data["admin"]["GPS_settings"]["location_text_edit"] + str(n))
        Logging("- Input edit location")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Save')]")))
        driver.find_element_by_xpath(data["admin"]["GPS_settings"]["button_save"]).click()
        driver.find_element_by_xpath(data["admin"]["GPS_settings"]["close_save"]).click()
        Logging("- Save")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["edit_GPS"]["pass"])
        time.sleep(10)
    except:
        Logging("- Can't edit GPS")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["edit_GPS"]["fail"])

    try:
        Logging("** Delete GPS")
        driver.swipe(start_x=1000, start_y=450, end_x=500, end_y=450, duration=800)

        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Delete')]"))).click()
        driver.find_element_by_xpath(data["admin"]["GPS_settings"]["delete"]).click()
        Logging("- Select GPS delete")
        driver.find_element_by_xpath(data["admin"]["GPS_settings"]["accept_delete"]).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Close')]"))).click()
        Logging("- Accept delete")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["delete_GPS"]["pass"])
        time.sleep(10)
    except:
        Logging("- Can't delete GPS")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["delete_GPS"]["fail"])

    driver.find_element_by_xpath(data["admin"]["GPS_settings"]["back"]).click()
    Logging(" ")

def admin_WF():
    Logging("------- Add Wifi -------")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'WiFi Setting')]"))).click()
    Logging("- Select Wifi setting")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["add_wifi"]).click()
    Logging("- Add new wifi")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["wifi_name"]).click()
    Logging("- Select name wifi")

    try:
        warning = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'WIFI Settings')]")))
        if warning.is_displayed():
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["admin"]["WIFI_settings"]["add_wifi"])))
        else:
            Logging("=> Fail")
    except WebDriverException:
        Logging("=> Fail")

    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["button_next"][0]).click()
    Logging("- Select next step")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["workPlace"]).click()
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["workPlace_input"]).click()
    Logging("- Select work Place")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["button_next"][1]).click()
    Logging("- Select next step")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["close_popup"]).click()
    Logging("- Close pop up")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["complete"]).click()
    Logging("- Add wifi success")
    time.sleep(5)

    '''
    * Can't use function delete when run auto appium
    -> When swpie element, delete button can't click on. 
    -> When finish auto, open app and delete element => can't click on button, must kill app and open again it's can be use again
    * Can delete when use manaul 
    '''
    Logging("Delete wifi")
    driver.swipe(start_x=1000, start_y=550, end_x=500, end_y=550, duration=800)
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Delete')]"))).click()
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["delete"]).click()
    Logging("- Select wifi")
    driver.find_element_by_xpath(data["admin"]["GPS_settings"]["accept_delete"]).click()
    driver.find_element_by_xpath(data["admin"]["GPS_settings"]["popup"]).click()
    Logging("- Accept delete")

    Logging("- Search wifi")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'WiFi Setting')]"))).click()
    Logging("- Select Wifi setting")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Please enter your keyword.')]"))).click()
    
    ''' Send key from keyboard mobile '''
    driver.is_keyboard_shown()
    driver.press_keycode(38)
    driver.press_keycode(47)
    driver.press_keycode(62)
    driver.press_keycode(50)
    driver.press_keycode(37)
    driver.press_keycode(42)
    driver.press_keycode(29)
    driver.press_keycode(66)

    driver.find_element_by_xpath(data["admin"]["GPS_settings"]["back"]).click()

def admin_BC():
    Logging(" ")
    Logging("------- Add Beacon -------")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Beacon Setting')]"))).click()
    Logging("- Select Beacon")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["add_wifi"]).click()
    Logging("- Add new Beacon")
    time.sleep(20)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'MiniBeacon_00997')]"))).click()
    Logging("- Select Beacon")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Next')]"))).click()
    Logging("- Select next step")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["workPlace"]).click()
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["workPlace_input"]).click()
    Logging("- Select work Place")
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["button_next"][1]).click()
    Logging("- Select next step")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Close')]"))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Complete')]"))).click()
    Logging("- Add Beacon success")
    driver.find_element_by_xpath(data["admin"]["GPS_settings"]["back"]).click()

    driver.find_element_by_xpath(data["admin"]["BEACON_settings"]["edit"]).click()
    beacon = driver.find_element_by_xpath(data["admin"]["BEACON_settings"]["edit_name"])
    beacon.clear()
    beacon.send_keys(data["admin"]["BEACON_settings"]["name"] + str(n))
    driver.find_element_by_xpath(data["admin"]["WIFI_settings"]["button_next"][1]).click()
    Logging("- Select next step")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Close')]"))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Complete')]"))).click()
    Logging("- Edit Beacon success")
    time.sleep(2)
    driver.find_element_by_xpath(data["admin"]["GPS_settings"]["back"]).click()

    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["admin"]["homepage"]))).click()

def settings():
    Logging(" ")
    Logging("** Settings app **")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["settings_button"]))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["language"]))).click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='한국어']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='닫기']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["settings_button"]))).click()
    
    korean_text = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["language_text"])))
    if korean_text.text == '한국어':
        Logging("=> Change to language '한국어' success")
    else:
        Logging("=> Fail")

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='언어 설정']"))).click()
    
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Tiếng Việt']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Đóng']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["settings_button"]))).click()

    VN_text = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["language_text"])))
    if VN_text.text == 'Tiếng Việt':
        Logging("=> Change to language 'Tiếng Việt' success")
    else:
        Logging("=> Fail")

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Thay đổi ngôn ngữ']"))).click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='日本語']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='閉じる']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["settings_button"]))).click()

    JP = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["language_text"])))
    if JP.text == '日本語':
        Logging("=> Change to language '日本語' success")
    else:
        Logging("=> Fail")

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='言語']"))).click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='简体中文']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='關閉']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["settings_button"]))).click()

    TQ = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["language_text"])))
    if TQ.text == '简体中文':
        Logging("=> Change to language '简体中文' success")
    else:
        Logging("=> Fail")

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='語言']"))).click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Indonesian']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Tutup']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["settings_button"]))).click()

    indo = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["language_text"])))
    if indo.text == 'Indonesian':
        Logging("=> Change to language 'Indonesian' success")
    else:
        Logging("=> Fail")

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Ganti BAHASA']"))).click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='English']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Close']"))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["settings_button"]))).click()
    EN = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["settings"]["language_text"])))
    if EN.text == 'English':
        Logging("=> Change to language 'English' success")
    else:
        Logging("=> Fail")

def attachfile():
    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["add_file"]))).click()
        Logging("- Select attach file")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["choose_photo"]))).click()
        Logging("- Choose photo")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, data["vacation"]["choose_gallery"]))).click()
        Logging("- Choose gallery")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["select_photo"]))).click()
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, data["vacation"]["select"]))).click()
        Logging("- Select photo")
        time.sleep(5)
        TesCase_LogResult(**data["testcase_result"]["vacation"]["attach_file"]["pass"])
    except:
        Logging("- Can't attach file")
        TesCase_LogResult(**data["testcase_result"]["vacation"]["attach_file"]["fail"])

def select_date_month():
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["next"]))).click()
    Logging("- View next date-month")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["prev"]))).click()
    Logging("- View pre date-month")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["calendar_select"]))).click()
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=19]"))).click()
    time.sleep(5)
    dateselect = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["date_calendar"])))
    dateselect_text = dateselect.text
    time.sleep(5)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["select_date"]))).click()
    Logging("- Select calendar")

    return dateselect_text

def input_user_name():
    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Please insert keyword to search']"))).click()

    ''' Send key "quynh" from keyboard mobile '''
    driver.is_keyboard_shown()
    driver.press_keycode(45)
    driver.press_keycode(49)
    driver.press_keycode(53)
    driver.press_keycode(42)
    driver.press_keycode(36)
    driver.press_keycode(66)
    Logging("- Search user")

def timecard():
    try:
        Logging(" ")
        Logging("** Check crash app **")
        Logging("- Timesheet - Daily -")
        driver.find_element_by_xpath(data["event"]["timecard"]).click()
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["timesheet"]))).click()

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_daily"]["next"]))).click()
        Logging("- View next date")   
        time.sleep(5)
        next_text = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_daily"]["policy"])))
        if next_text.text == 'Work Policy':
            Logging("=> View next date success")
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)    

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_daily"]["prev"]))).click()
        Logging("- View preview date")
        time.sleep(5)
        prev_text = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_daily"]["policy"])))
        if prev_text.text == 'Work Policy':
            Logging("=> View prev date success")
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)   

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_daily"]["calendar"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=9]"))).click()
        Logging("- Select date from calendar")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_daily"]["select_date"]))).click()
        time.sleep(5)
        date_select = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_daily"]["date"])))
        if date_select.is_displayed():
            Logging("=> Select date success")
        else:
            Logging("=> Crash app")
            exit(0)   
        time.sleep(5)   
        
        Logging(" ")
        Logging("- Timesheet - List -")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["list"]))).click()
        Logging("- Go to List") 
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["next"]))).click()
        Logging("- View next month")   
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["prev"]))).click()
        Logging("- View preview month")
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["calendar"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=9]"))).click()
        Logging("- Select date from calendar")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["select_date"]))).click()
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["sort"]))).click()
        Logging("- Sort by")
        time.sleep(2)
        list_week = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["list_sort"])))
        if list_week.is_displayed():
            Logging("- Show list week")
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_2"]))).click()
        Logging("- 2nd Week")
        time.sleep(5)
        
        total_week_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_2_text"])))
        if total_week_1.text == 'TOTAL OF 2ND WEEK':
            Logging("=> TOTAL OF 2ND WEEK")
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["sort"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_3"]))).click()
        Logging("- 3rd Week")
        time.sleep(5)

        total_week_2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_3_text"])))
        if total_week_2.text == 'TOTAL OF 3RD WEEK':
            Logging("=> TOTAL OF 3RD WEEK")
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["sort"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_4"]))).click()
        Logging("- 4th Week")
        time.sleep(5)
        total_week_3 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_4_text"])))
        if total_week_3.text == 'TOTAL OF 4TH WEEK':
            Logging("=> TOTAL OF 4TH WEEK")
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["sort"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_5"]))).click()
        Logging("- 5th Week")
        time.sleep(5)
        total_week_4 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["week_5_text"])))
        if total_week_4.text == 'TOTAL OF 5TH WEEK':
            Logging("=> TOTAL OF 5TH WEEK")
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)
        
        Logging(" ")
        Logging("- Timesheet - Calendar -")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["calendar"]))).click()
        Logging("- Go to List") 
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["next"]))).click()
        Logging("- View next date")   
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["prev"]))).click()
        Logging("- View preview date")
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["calendar_select"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=9]"))).click()
        Logging("- Select date from calendar")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["select_date"]))).click()
        time.sleep(5)
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_Timesheet"]["pass"])
    except:
        Logging("-> Crash app")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_Timesheet"]["fail"])
    
    driver.find_element_by_xpath(data["event"]["timecard"]).click()

    ''' REPORT '''
    try:
        Logging("** Check report - Monthly")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["MT_report"]))).click()
        Logging("- Schedule Working")
        time.sleep(10)

        schedule = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["schedule_working"])))
        if schedule.text == 'Scheduled working day':
            count_day = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["count_schedule_working"])))
            Logging("- Scheduled working day:", count_day.text)
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["events"]))).click()
        Logging("- Events")
        time.sleep(5)
        clock_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["clockin"])))
        if clock_in.text == 'Clock-In':
            count_clock_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["count_clockin"])))
            Logging("- Events - Clock in:", count_clock_in.text)
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["working_status"]))).click()
        Logging("- Working status")
        time.sleep(5)
        working_time = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["workingtime"])))
        if working_time.text == 'Working time':
            count_working_time = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_monthly"]["count_workingtime"])))
            Logging("- Working status - Working time:", count_working_time.text)
        else:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        Logging("** Check report - Weekly")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_weekly"]["weekly"]))).click()
        Logging("- View week tab")
        time.sleep(10)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Device']"))).click()
        Logging("- View tab device")
        time.sleep(5)

        driver.swipe(start_x=1174, start_y=730, end_x=400, end_y=730, duration=800)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Average working hour per week']"))).click()
        Logging("- View tab Avg_Working")
        time.sleep(5)
        driver.swipe(start_x=1174, start_y=730, end_x=400, end_y=730, duration=800)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Working hours per day of the week']"))).click()
        Logging("- View tab working hour")
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["next"]))).click()
        Logging("- View next date")   
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["prev"]))).click()
        Logging("- View preview date")
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["calendar_select"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=9]"))).click()
        Logging("- Select date from calendar")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["select_date"]))).click()
        time.sleep(5)

        Logging("** Check report - List")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report_list"]["list"]))).click()
        Logging("- View list tab")
        time.sleep(10)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["next"]))).click()
        Logging("- View next date")   
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["prev"]))).click()
        Logging("- View preview date")
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_calendar"]["calendar_select"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=9]"))).click()
        Logging("- Select date from calendar")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timesheet_list"]["select_date"]))).click()
        time.sleep(5)
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_Report"]["pass"])
    except:
        Logging("-> Crash app")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_Report"]["fail"])

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,data["event"]["timecard"]))).click()
    
    try:
        Logging(" ")
        Logging("** Check Company timecard - Weekly Status")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["weekly_status_CT"]))).click()
        time.sleep(5)
        try:
            title = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["week_title"])))
            if title.text == 'Weekly Status':
                Logging("- Show content")

        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["sort_week"]))).click()
            Logging("- Select sort week")
            sortweek = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["list"])))
            if sortweek.is_displayed:
                Logging("- List week display")
            else:
                Logging("- List not display")
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        Logging("- Select week")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["week_2"]))).click()
        Logging("- 2nd Week")
        time.sleep(5)
        try:
            user_view = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["user"])))
            if user_view.is_displayed():
                Logging("- Show content")
            else:
                Logging("- No data")
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["sort_week"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["week_3"]))).click()
        Logging("- 3rd Week")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["sort_week"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["week_4"]))).click()
        Logging("- 4th Week")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["sort_week"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["week_5"]))).click()
        Logging("- 5th Week")
        time.sleep(5)

        '''WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["sort_week"]))).click()
        time.sleep(5)
        driver.swipe(start_x=291, start_y=1126, end_x=291, end_y=900, duration=800)
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["week_6"]))).click()
        Logging("- 6th Week")
        time.sleep(5)'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["sort_week"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["week_1"]))).click()
        Logging("- 1st Week")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["search"]))).click()
        input_user_name()

        WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'quynh1')]"))).click()
        time.sleep(5)

        ''' Next - Prev - Select date '''
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["next"]))).click()
        Logging("- View next date-month")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["prev"]))).click()
        Logging("- View pre date-month")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["calendar_select"]))).click()
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=19]"))).click()
        time.sleep(5)
        dateselect = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["date_calendar"])))
        dateselect_text = dateselect.text
        time.sleep(5)
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["select_date"]))).click()
        Logging("- Select calendar")

        time.sleep(10)

        try:
            user_view = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["user"])))
            if user_view.is_displayed():
                Logging("- Show content")
            else:
                Logging("- No data")
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["view_detail"]))).click()
        Logging("- View detail")

        date_of_calendar = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["date"])))
        date = date_of_calendar.text
        x = date.split(" ")[2]
        a = x.split(",")[0]

        if a == dateselect_text:
            Logging("- Show right date")
        else:
            Logging("- Crash")
            exit(0)
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_weekly_status"]["pass"])
    except:
        Logging("-> Crash app")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_weekly_status"]["fail"])

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["weekly_status"]["back"]))).click()

    try:
        Logging(" ")
        Logging("** Check Company timecard - Time Line")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["timeline_CT"]))).click()
        time.sleep(5)
        try:
            title = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["timeline_title"])))
            if title.text == 'Time Line':
                Logging("- Show content")

        except WebDriverException:
            Logging("=> Crash app")
            exit(0)

        ''' Search user '''
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["search"]))).click()
        input_user_name()

        WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'quynh1')]"))).click()
        Logging("- Select user")
        time.sleep(5)

        ''' Next - Prev - Select date '''
        select_date_month()
        time.sleep(10)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        Logging("- Sort time line")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["clockin"]))).click()
        Logging("- Clock in")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        Logging("- Sort time line")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["tardin"]))).click()
        Logging("- Tardines")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        Logging("- Sort time line")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["clockout"]))).click()
        Logging("- Clock out")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        Logging("- Sort time line")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["auto_clockout"]))).click()
        Logging("- Automatically Clock-out")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        Logging("- Sort time line")
        time.sleep(5)
        driver.swipe(start_x=291, start_y=1315, end_x=291, end_y=691, duration=800)
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["outside"]))).click()
        Logging("- Outside")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        time.sleep(5)
        driver.swipe(start_x=291, start_y=1315, end_x=291, end_y=691, duration=800)
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["meeting"]))).click()
        Logging("- Meeting")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        time.sleep(5)
        driver.swipe(start_x=291, start_y=1315, end_x=291, end_y=691, duration=800)
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["edu"]))).click()
        Logging("- Education")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        time.sleep(5)
        driver.swipe(start_x=291, start_y=1315, end_x=291, end_y=691, duration=800)
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["business"]))).click()
        Logging("- Business Trip")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        time.sleep(5)
        driver.swipe(start_x=291, start_y=1315, end_x=291, end_y=691, duration=800)
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["working_remote"]))).click()
        Logging("- Working remote")
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["sort_timeline"]))).click()
        time.sleep(5)
        driver.swipe(start_x=291, start_y=1315, end_x=291, end_y=691, duration=800)
        driver.swipe(start_x=291, start_y=1315, end_x=291, end_y=691, duration=800)
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["timeline"]["working"]))).click()
        Logging("- Working")
        time.sleep(5)
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_time_line"]["pass"])
    except:
        Logging("-> Crash app")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_time_line"]["fail"])

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,data["event"]["timecard"]))).click() 

    try:
        Logging(" ")
        Logging("** Check Company timecard - Report")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["report_CT"]))).click()
        Logging("- View by work")
        time.sleep(5)
        try:
            title = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["report_title"])))
            if title.text == 'Report':
                Logging("- Show content")
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)

        ''' Next - Prev - Select date '''
        select_date_month()
        time.sleep(10)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["search"]))).click()
        Logging("- Search user")
        input_user_name()
        time.sleep(3)
        WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'quynh1')]"))).click()
        Logging("- Select user")
        time.sleep(5)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["detail"]))).click()
        Logging("- View detail")
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["back"]))).click()
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["view_by_event"]))).click()
        Logging("- View by event")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["detail"]))).click()
        Logging("- View detail")
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["back"]))).click()
        time.sleep(5)
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_report"]["pass"])
    except:
        Logging("-> Crash app")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_report"]["fail"])

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,data["event"]["timecard"]))).click()

    try:
        Logging(" ")
        Logging("** Check Company timecard - Approval")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approval_CT"]))).click()
        time.sleep(5)
        try:
            title = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approval_title"])))
            if title.text == 'Approval':
                Logging("- Show content")
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        time.sleep(10)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["filter_type"]))).click()
        Logging("- Filter type")
        time.sleep(5)
        try:
            filter = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["detail"])))
            if filter.text == 'Detail':
                Logging("-")
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["type"]))).click()
        Logging("- Select type")
        time.sleep(5)
        try:
            type_detail = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["event"])))
            if type_detail.is_displayed():
                type_detail.click()
                time.sleep(5)
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["status"]))).click()
        Logging("- Select status")
        time.sleep(5)
        try:
            status = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["pending"])))
            if status.is_displayed():
                status.click()
                time.sleep(5)
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)

        '''WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["peroid"]))).click()
        Logging("- Select peroid")
        time.sleep(5)
        try:
            peroid = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["today"])))
            if peroid.is_displayed():
                peroid.click()
                time.sleep(5)
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["back"]))).click()
        time.sleep(5)
        try:
            title = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approval_title"])))
            if title.text == 'Approval':
                Logging("- Show content")
        except WebDriverException:
            Logging("=> Crash app")
            exit(0)
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["report"]["search"]))).click()
        
        input_user_name()

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["user"]))).click()
        #WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text='quynh1']"))).click()
        Logging("- Select user")
        time.sleep(5)

        ''' Approve '''
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["select"]))).click()
        try:
            approve_line = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approval_line"])))
            if approve_line.is_displayed():
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approve"]))).click()
                time.sleep(5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["close"]))).click()
                Logging("- Approve request")
                time.sleep(5)
                
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["filter_type"]))).click()
                Logging("- Filter type")
                time.sleep(5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["status"]))).click()
                Logging("- Select status")
                time.sleep(5)
                WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approved"]))).click()
                time.sleep(5)

                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["back"]))).click()
                time.sleep(5)

                status_approve = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["status_text"])))
                if status_approve.text == 'Approved':
                    Logging("- Approve success")
                else:
                    Logging("- Fail")
            else:
                Logging("=> Approve don't have approve permission")
        except WebDriverException:
            Logging("=> Approve don't have approve permission")

        ''' Reject '''
        add_event()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["event"]["timecard"]))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approval_CT"]))).click()
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["select"]))).click()
        try:
            approve_line = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["approval_line"])))
            if approve_line.is_displayed():
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["reject_bt"]))).click()
                time.sleep(5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["close"]))).click()
                time.sleep(5)
                status_reject = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TimeCard"]["approval"]["status_text"])))
                if status_reject.text == 'Rejected':
                    Logging("- Rejected success")
                else:
                    Logging("- Fail")
            else:
                Logging("=> Approve don't have approve permission")
        except WebDriverException:
            Logging("=> Approve don't have approve permission")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_Approval"]["pass"])
    except:
        Logging("-> Crash app")
        TesCase_LogResult(**data["testcase_result"]["timecard"]["crash_app_Approval"]["pass"])

def request_vacation():
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["button_vacation"]))).click()
    Logging("- Vacation")
    ''' Request vaction'''
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["request"]))).click() 
    
    title_request = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["request_vacation_text"])))
    if title_request.text == 'Request vacation':
        Logging("=> Request vacation")
    else:
        Logging("=> Crash app")
        exit(0)   
    try:
        vacation_type = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["AM"])))
        if vacation_type.is_displayed():
            vacation_type.click()
            Logging("- Select vacation type")

            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["calendar"]))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=16]"))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=16]"))).click()
            time.sleep(2)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["select_calendar"]))).click()  
            time.sleep(2)

            ''' Crash app when select date '''
            try:
                title_request = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["request_vacation_text"])))
                if title_request.text == 'Request vacation':
                    Logging("- Select date vacation")
                else:
                    Logging("=> Crash app")
                    exit(0)  
            except WebDriverException: 
                Logging("=> Crash app")
                exit(0)

            ''' Get data of vacation request '''
            vacation_date = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["request_date_text"])))
            vacation_text = vacation_date.text
            date_text = vacation_text.split(" ")[0]
            type_vacation = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["AM"])))
            type_text = type_vacation.text
            vacation_date_type = date_text + "(" + type_text + ")"            
    except:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["calendar"]))).click()
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=16]"))).click()
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=16]"))).click()
        time.sleep(2)
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["select_calendar"]))).click()  
        time.sleep(2)

        ''' Get data of vacation request '''
        vacation_date = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["request_date_text"])))
        vacation_text = vacation_date.text
        # date_text = vacation_text.split(" ")[0]
        # type_vacation = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["AM"])))
        # type_text = type_vacation.text
        # vacation_date_type = vacation_text + "(" + type_text + ")"

    # Attachment
    try:
        attachfile()
    except:
        pass

    driver.swipe(start_x=650, start_y=1844, end_x=650, end_y=355, duration=800)
    time.sleep(5)
    ''' Select CC '''
    # WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["CC"]))).click()
    # input_user_name()
    # WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'quynh1')]"))).click()
    # WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["save_cc"]))).click()
    # Logging("- Select CC")

    driver.swipe(start_x=650, start_y=1662, end_x=650, end_y=355, duration=800)

    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Please enter your reason']"))).send_keys(data["vacation"]["my_vacation"]["input_test"])
    Logging("- Input reason")
    
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["button_request"]))).click()
    
    '''- Check day request
      + If vacation day is saturday => fail, check again
      + If memo is empty => fail, check again'''
    try:
        fail = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["request_fail"])))
        if fail.text == 'request vacation failure':
            Logging("--- Request vacation failure - vacation day is saturday---")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["close_fail"]))).click()
            time.sleep(2)
            driver.swipe(start_x=650, start_y=355, end_x=650, end_y=2275, duration=800)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["calendar"]))).click()
            time.sleep(2)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=16]"))).click()
            Logging("=> Select start date")
            
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.view.ViewGroup[@index='1']//android.widget.Button[@index=16]"))).click()
            Logging("=> Select end date")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["select_calendar"]))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["button_request"]))).click() 
            Logging("=> Send request vacation")
            TesCase_LogResult(**data["testcase_result"]["vacation"]["request_vacation"]["pass"])
        else:
            Logging("=> Request success")
            TesCase_LogResult(**data["testcase_result"]["vacation"]["request_vacation"]["pass"])
    except WebDriverException:
        Logging("=> Request success") 
        TesCase_LogResult(**data["testcase_result"]["vacation"]["request_vacation"]["pass"])

    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["button_close"]))).click()

def approve_request():
    '''
    --- User 1 ---
    1. Request vacation
    2. Send request to user 2
    Log out account user 1

    --- User 2 ---
    Log in to account user 2
    => Approve request

    --- User 1 ---
    Log in to account user 1
    => Check request vacation
    '''
    ''' Case of user 1 '''
    time.sleep(15)
    request_vacation()

    Logging(" ")
    ''' Log out '''
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["setting_button"]))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["logout"]))).click()
    Logging("=> Change to user 2 to approve request")

    ''' Log in '''
    id_user2 = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["user_name"])))
    id_user2.clear()
    id_user2.send_keys(data["id_name_2"])
    Logging("- Input ID")
    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Password']")))
    driver.find_element_by_xpath(data["password"]).send_keys(data["pass_input"])
    Logging("- Input Password")
    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Login')]")))
    driver.find_element_by_xpath(data["button_login"]).click()
    Logging("=> Click Log In button")
    driver.implicitly_wait(50)

    ''' Check request vacation of user 1 '''
    time.sleep(5)
    Logging("- Check request vacation")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["button_vacation"]))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["manage_processing"]["vacation_approve"]))).click()
    time.sleep(3)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["search"]))).click()
    input_user_name()
    Logging("- Search user")

    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'quynh1')]"))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["select_user"]))).click()
    Logging("- Select user")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["approve"]))).click()
    Logging("- Approve request")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["accept_approve"]))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["close_popup"]))).click()
    Logging("=> Approve success")

    text_approve = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["approve_text"])))
    if text_approve.text == 'Approved':
        Logging("=> Request have approve success")
        TesCase_LogResult(**data["testcase_result"]["vacation"]["approve_request"]["pass"])
    else:
        Logging("=> Approve fail")
        TesCase_LogResult(**data["testcase_result"]["vacation"]["approve_request"]["fail"])

    ''' User cancel request '''
    Logging(" ")
    ''' Log out '''
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["setting_button"]))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["logout"]))).click()
    Logging("=> Change to user 1 - check request have been approve - cancel request")

    ''' Log in '''
    id_user2 = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["user_name"])))
    id_user2.clear()
    id_user2.send_keys(data["id_name"])
    Logging("- Input ID")
    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Password']")))
    driver.find_element_by_xpath(data["password"]).send_keys(data["pass_input"])
    Logging("- Input Password")
    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Login')]")))
    driver.find_element_by_xpath(data["button_login"]).click()
    Logging("=> Click Log In button")
    driver.implicitly_wait(50)

    ''' Check request vacation of user 1 '''
    time.sleep(5)
    Logging("- Check request vacation")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["button_vacation"]))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["manage_processing"]["vacation_approve"]))).click()
    time.sleep(3)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["search"]))).click()
    input_user_name()

    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'quynh1')]"))).click()

    text_approve = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["approve_text"])))
    if text_approve.text == 'Approved':
        Logging("=> Request have approve success")
        TesCase_LogResult(**data["testcase_result"]["vacation"]["approve_request"]["pass"])
    else:
        Logging("=> Approve fail")
        TesCase_LogResult(**data["testcase_result"]["vacation"]["approve_request"]["fail"])

    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["button_vacation"]))).click()
    time.sleep(5)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["vacation_status"]["vacationstatus"]))).click()
    time.sleep(10)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["vacation_status"]["cancel_request"]))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["vacation_status"]["button_ok"]))).click()
    Logging("- Cancel request")
    time.sleep(10)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["close_popup"]))).click()
    Logging("=> Approve cancel request success")
    time.sleep(5)

    text_cancel = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["my_vacation"]["vacation_status"]["text_request"])))
    if text_cancel.text == 'User cancel':
        Logging("=> Send cancel request success")
    else:
        Logging("=> Approve Arbitrary decision")

    Logging(" ")
    ''' Change to user 2 - approve cancel request'''
    ''' Log out '''
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["setting_button"]))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["logout"]))).click()
    Logging("=> Change to user 2 - approve cancel request")

    ''' Log in '''
    id_user2 = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["user_name"])))
    id_user2.clear()
    id_user2.send_keys(data["id_name_2"])
    Logging("- Input ID")
    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Password']")))
    driver.find_element_by_xpath(data["password"]).send_keys(data["pass_input"])
    Logging("- Input Password")
    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@text,'Login')]")))
    driver.find_element_by_xpath(data["button_login"]).click()
    Logging("=> Click Log In button")
    driver.implicitly_wait(50)

    ''' Check request vacation of user 1 '''
    time.sleep(5)
    try:
        Logging("- Check request vacation")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["button_vacation"]))).click()
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["manage_processing"]["vacation_approve"]))).click()
        time.sleep(3)
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["cancel_request"]))).click()
        Logging("- Click cancel request")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["status"]))).click()
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//*[@text='Request']"))).click()
        Logging("- Select status request")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["request"]))).click()
        Logging("- Select request")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["approve_cancel"]))).click()
        Logging("- APPROVE CANCELLATION")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["accept_approve_cancel"]))).click()
        Logging("=> Cancel request")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["vacation"]["vacation_approve"]["close_popup"]))).click()
        TesCase_LogResult(**data["testcase_result"]["vacation"]["approve_cancel"]["pass"])
    except:
        Logging("=> Cancel request fail")
        TesCase_LogResult(**data["testcase_result"]["vacation"]["approve_cancel"]["fail"])

def check_notification():
    request_vacation()
    time.sleep(10)
    driver.swipe(start_x=713, start_y=0, end_x=713, end_y=1189, duration=800)
    Logging("- Sroll action bar")
    try:
        noti_text = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.widget.FrameLayout[1]/android.view.ViewGroup/android.widget.ScrollView/android.widget.FrameLayout[1]/android.widget.FrameLayout[1]/android.widget.FrameLayout//android.widget.LinearLayout[1]/android.view.ViewGroup/android.widget.TextView[1]")))
        vacation_text = noti_text.text
        Logging(vacation_text)
        user_text = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//android.widget.FrameLayout[1]/android.view.ViewGroup/android.widget.ScrollView/android.widget.FrameLayout[1]/android.widget.FrameLayout[1]/android.widget.FrameLayout//android.widget.LinearLayout[2]/android.widget.FrameLayout/android.widget.TextView")))
        user = user_text.text
        Logging(user)
        type_text = vacation_text.split(" ")[0]
        if type_text == "[Request]":
            Logging("- Notification have show")
    except:
        Logging("- Don't show notification")

